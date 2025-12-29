# -*- coding: utf-8 -*-
"""
Excel Yardımcı Modülü
Excel dosyalarını okuma, yazma ve formatlama işlemleri.
"""

import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill

# Config'i import et
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import EXCEL_STIL, EXCEL_AYARLARI


class ExcelYardimci:
    """Excel işlemleri için yardımcı sınıf"""
    
    @staticmethod
    def format_sure(delta):
        """
        Timedelta'yı saat:dakika:saniye olarak biçimlendir.
        
        Args:
            delta: timedelta objesi
            
        Returns:
            str: "HH:MM:SS" formatında süre
        """
        total_seconds = int(delta.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    @staticmethod
    def format_tarih(val):
        """
        Tarih/saat değerini gg.aa.yyyy ss:dd:nn formatına çevir.
        
        Args:
            val: Tarih değeri
            
        Returns:
            str: Formatlanmış tarih
        """
        if val == '' or val is None:
            return ''
        
        val_str = str(val)
        
        try:
            if hasattr(val, 'strftime'):
                return val.strftime('%d.%m.%Y %H:%M:%S')
        except:
            pass
        
        try:
            if '(' in val_str:
                parts = val_str.split('(')
                date_part = parts[0].strip()
                extra = f" ({parts[1]}" if len(parts) > 1 else ""
                if '-' in date_part and ':' in date_part:
                    dt = pd.to_datetime(date_part)
                    return dt.strftime('%d.%m.%Y %H:%M:%S') + extra
            elif '-' in val_str and ':' in val_str and len(val_str) > 10:
                dt = pd.to_datetime(val_str)
                return dt.strftime('%d.%m.%Y %H:%M:%S')
        except:
            pass
        
        return val_str
    
    @staticmethod
    def temizle_ve_formatla(val, max_karakter=None, wrap_satir=False):
        """
        Değeri temizle ve formatla (PNG için optimize edilmiş).
        
        Args:
            val: Formatlanacak değer
            max_karakter: Maksimum karakter sayısı (None ise sınırsız)
            wrap_satir: Uzun metinleri satırlara böl
            
        Returns:
            str: Temizlenmiş ve formatlanmış değer
        """
        if val == '' or val is None or pd.isna(val):
            return ''
        
        # Önce tarih formatına çevir
        formatted = ExcelYardimci.format_tarih(val)

        # Sayısal değerleri temizle / yuvarla (özellikle süre hücreleri için)
        # Örnek: 12.345678 → 12.35
        try:
            num_val = float(str(formatted).replace(',', '.'))
            # 2 ondalık basamağa yuvarla
            formatted = f"{num_val:.2f}"
            # 12.00 → 12, 12.30 → 12.3
            if '.' in formatted:
                formatted = formatted.rstrip('0').rstrip('.')
        except Exception:
            # Sayısal değilse olduğu gibi devam et
            pass
        
        # String temizleme
        if isinstance(formatted, str):
            # Gereksiz boşlukları temizle
            formatted = ' '.join(formatted.split())
            
            # NaN, None gibi değerleri temizle
            if formatted.lower() in ['nan', 'none', 'null', '']:
                return ''
        
        # Maksimum karakter kontrolü
        if max_karakter and len(formatted) > max_karakter:
            if wrap_satir:
                # Akıllı satır bölme
                words = formatted.split()
                lines = []
                current_line = ""
                for word in words:
                    if len(current_line + " " + word) <= max_karakter:
                        current_line += (" " if current_line else "") + word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word
                        if len(word) > max_karakter:
                            # Çok uzun kelimeyi böl
                            while len(word) > max_karakter:
                                lines.append(word[:max_karakter])
                                word = word[max_karakter:]
                            current_line = word
                if current_line:
                    lines.append(current_line)
                formatted = "\n".join(lines)
            else:
                formatted = formatted[:max_karakter-3] + "..."
        
        return formatted
    
    @staticmethod
    def normalize_grup_string(grup_str):
        """
        Grup string'ini normalize et (noktalı virgülle ayrılmış ID'ler).
        
        Args:
            grup_str: Grup string'i
            
        Returns:
            str: Normalize edilmiş string
        """
        grup_str = str(grup_str).strip()
        
        if ';' in grup_str:
            parts = []
            for part in grup_str.split(';'):
                part = part.strip()
                try:
                    if '.' in part:
                        part = str(int(float(part)))
                    else:
                        part = str(int(part))
                except:
                    pass
                parts.append(part)
            return ';'.join(parts)
        else:
            try:
                if '.' in grup_str:
                    return str(int(float(grup_str)))
                else:
                    return str(int(grup_str))
            except:
                return grup_str
    
    @staticmethod
    def kaydet_bicimli(df, dosya_yolu):
        """
        DataFrame'i formatlı Excel olarak kaydet.
        
        Args:
            df: Kaydedilecek DataFrame
            dosya_yolu: Hedef dosya yolu
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Birlesik_Analiz"
        
        # Başlıkları yaz
        ws.append(list(df.columns))
        
        # Verileri yaz
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Hücre formatlaması
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Sütun genişliklerini ayarla
        max_width = EXCEL_STIL.get('MAX_COLUMN_WIDTH', 60)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)
        
        wb.save(dosya_yolu)
        print(f"✓ Excel kaydedildi: {dosya_yolu}")
    
    @staticmethod
    def kaydet_stillendirilmis(df, dosya_yolu, sheet_adi="Veri"):
        """
        DataFrame'i stillendirilmiş Excel olarak kaydet.
        
        Args:
            df: Kaydedilecek DataFrame
            dosya_yolu: Hedef dosya yolu
            sheet_adi: Sayfa adı
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_adi[:31]  # Excel max 31 karakter
        
        header_fill = PatternFill(
            start_color=EXCEL_STIL['HEADER_BG_COLOR'], 
            end_color=EXCEL_STIL['HEADER_BG_COLOR'], 
            fill_type="solid"
        )
        header_font = Font(bold=True, color=EXCEL_STIL['HEADER_FONT_COLOR'])
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_fill_even = PatternFill(
            start_color=EXCEL_STIL['ALTERNATE_ROW_COLOR'], 
            end_color=EXCEL_STIL['ALTERNATE_ROW_COLOR'], 
            fill_type="solid"
        )
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Başlıkları yaz
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Verileri yaz
        for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(data_row.values, 1):
                cell_value = ExcelYardimci.format_tarih(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = data_alignment
                if row_idx % 2 == 0:
                    cell.fill = data_fill_even
        
        # Sütun genişliklerini ayarla
        max_width = EXCEL_STIL.get('MAX_COLUMN_WIDTH', 50)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(dosya_yolu)
        print(f"✓ Excel kaydedildi: {dosya_yolu}")
    
    @staticmethod
    def oku_kesinti_dosyasi(dosya_yolu):
        """
        Kesinti Excel dosyasını oku.
        
        Args:
            dosya_yolu: Dosya yolu
            
        Returns:
            DataFrame veya None
        """
        try:
            df = pd.read_excel(
                dosya_yolu, 
                header=EXCEL_AYARLARI['KESINTI_HEADER_ROW']
            )
            print(f"✓ Kesinti dosyası okundu: {len(df)} satır")
            return df
        except Exception as e:
            print(f"✗ Dosya okunamadı: {e}")
            return None
    
    @staticmethod
    def oku_veri_dosyasi(dosya_yolu):
        """
        veri.xlsx / Birlesik_Analiz dosyasını oku.
        
        Args:
            dosya_yolu: Dosya yolu
            
        Returns:
            DataFrame veya None
        """
        try:
            df = pd.read_excel(
                dosya_yolu, 
                header=EXCEL_AYARLARI['VERI_HEADER_ROW'],
                keep_default_na=False
            )
            print(f"✓ Veri dosyası okundu: {len(df)} satır")
            return df
        except Exception as e:
            print(f"✗ Dosya okunamadı: {e}")
            return None
    
    @staticmethod
    def id_ara(df, aranan_id, col_index):
        """
        DataFrame'de ID ara.
        
        Args:
            df: Aranacak DataFrame
            aranan_id: Aranacak ID
            col_index: Aranacak sütun indeksi
            
        Returns:
            DataFrame: Eşleşen satırlar
        """
        try:
            aranan_str = str(aranan_id).strip()
            search_col = df.iloc[:, col_index]
            
            # String olarak ara
            eslesen = df[search_col.astype(str).str.strip() == aranan_str]
            if len(eslesen) > 0:
                return eslesen
            
            # Float olarak ara
            try:
                aranan_float = float(aranan_id)
                eslesen = df[search_col == aranan_float]
                if len(eslesen) > 0:
                    return eslesen
            except:
                pass
            
            # Int olarak ara
            try:
                aranan_int = int(float(aranan_id))
                eslesen = df[search_col == aranan_int]
                if len(eslesen) > 0:
                    return eslesen
            except:
                pass
            
            # Numeric dönüşüm ile ara
            try:
                aranan_float = float(aranan_id)
                search_col_numeric = pd.to_numeric(search_col, errors='coerce')
                eslesen = df[search_col_numeric == aranan_float]
                if len(eslesen) > 0:
                    return eslesen
            except:
                pass
            
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Arama hatası: {e}")
            return pd.DataFrame()

