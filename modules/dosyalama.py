# -*- coding: utf-8 -*-
"""
Dosyalama Modülü
PNG ve Excel raporları oluşturma işlemleri.
NOT: Ortak W değerleri hesaplanmaz, analiz sonucundan alınır.
"""

import pandas as pd
import os
import sys
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Config ve diğer modülleri import et
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    TABLE_SUTUN_INDEKSLERI, JTK_SUTUN_INDEKSLERI, 
    CM_SUTUN_INDEKSLERI, VERI_SUTUN_INDEKSLERI,
    PNG_AYARLARI, EXCEL_STIL, EXCEL_AYARLARI, VARSAYILAN,
    DAGITIM_AG_AYARLARI
)
from modules.excel_yardimci import ExcelYardimci


class Dosyalama:
    """Dosyalama ve raporlama işlemleri için sınıf"""
    
    def __init__(self, klasor_yolu):
        """
        Dosyalama sınıfını başlat.
        
        Args:
            klasor_yolu: Çalışma klasörü
        """
        self.klasor_yolu = klasor_yolu
        self.grup_list = []
        self.df_table = None
        self.df_jtk = None
        self.df_cm = None
        self.df_analiz = None  # Analiz sonucu (W değerleri dahil)
    
    def dosyalari_yukle(self):
        """
        Gerekli dosyaları yükle.
        
        Returns:
            tuple: (başarılı, eksik dosyalar listesi)
        """
        eksik = []
        
        table_path = os.path.join(self.klasor_yolu, 'table.xlsx')
        jtk_path = os.path.join(self.klasor_yolu, 'jtk.xlsx')
        cm_path = os.path.join(self.klasor_yolu, 'cm.xlsx')
        
        if not os.path.exists(table_path):
            eksik.append('table.xlsx')
        if not os.path.exists(jtk_path):
            eksik.append('jtk.xlsx')
        if not os.path.exists(cm_path):
            eksik.append('cm.xlsx')
        
        if eksik:
            return False, eksik
        
        try:
            self.df_table = pd.read_excel(
                table_path, 
                skiprows=EXCEL_AYARLARI['TABLE_SKIP_ROWS'], 
                header=0, 
                keep_default_na=False
            )
            self.df_jtk = pd.read_excel(
                jtk_path, 
                header=EXCEL_AYARLARI['JTK_HEADER_ROW'], 
                keep_default_na=False
            )
            self.df_cm = pd.read_excel(
                cm_path, 
                skiprows=2, 
                header=0, 
                keep_default_na=False
            )
            print(f"✓ Dosyalar yüklendi")
            return True, []
        except Exception as e:
            print(f"✗ Dosya yükleme hatası: {e}")
            return False, [str(e)]
    
    def analiz_sonucunu_yukle(self, analiz_yolu):
        """
        Analiz sonuç dosyasını yükle (W değerleri dahil).
        
        Args:
            analiz_yolu: Birlesik_Analiz.xlsx yolu
            
        Returns:
            bool: Başarılı ise True
        """
        try:
            self.df_analiz = pd.read_excel(analiz_yolu, header=0, keep_default_na=False)
            print(f"✓ Analiz sonucu yüklendi: {len(self.df_analiz)} satır")
            return True
        except Exception as e:
            print(f"✗ Analiz yüklenemedi: {e}")
            return False
    
    def gruplari_yukle(self, veri_yolu=None):
        """
        Grupları yükle (analiz sonucundan veya veri.xlsx'den).
        
        Args:
            veri_yolu: Opsiyonel veri dosyası yolu
            
        Returns:
            list: Grup listesi
        """
        self.grup_list = []
        
        # Önce analiz sonucundan dene
        if self.df_analiz is not None:
            try:
                h_sutun_idx = VERI_SUTUN_INDEKSLERI['ILGILI_KESINTILER']
                for idx, row in self.df_analiz.iterrows():
                    grup_str = str(row.iloc[h_sutun_idx]).strip()
                    if grup_str and grup_str != '':
                        grup_normalized = ExcelYardimci.normalize_grup_string(grup_str)
                        if grup_normalized not in self.grup_list:
                            self.grup_list.append(grup_normalized)
                print(f"✓ Analiz sonucundan {len(self.grup_list)} grup yüklendi")
                return self.grup_list
            except Exception as e:
                print(f"Analiz'den grup yükleme hatası: {e}")
        
        # veri.xlsx'den yükle
        if veri_yolu and os.path.exists(veri_yolu):
            try:
                df_veri = pd.read_excel(veri_yolu, header=0, keep_default_na=False)
                h_sutun_idx = VERI_SUTUN_INDEKSLERI['ILGILI_KESINTILER']
                
                for idx in range(1, len(df_veri)):
                    grup_str = str(df_veri.iloc[idx, h_sutun_idx]).strip()
                    if grup_str and grup_str != '':
                        grup_normalized = ExcelYardimci.normalize_grup_string(grup_str)
                        if grup_normalized not in self.grup_list:
                            self.grup_list.append(grup_normalized)
                
                print(f"✓ veri.xlsx'den {len(self.grup_list)} grup yüklendi")
            except Exception as e:
                print(f"veri.xlsx'den grup yükleme hatası: {e}")
        
        return self.grup_list
    
    def grup_icin_w_degerini_al(self, grup):
        """
        Analiz sonucundan grubun W değerini al (HESAPLAMA YAPMAZ).
        
        Args:
            grup: Grup string'i (;'li kesinti ID'leri)
            
        Returns:
            str: W değeri veya boş string
        """
        if self.df_analiz is None:
            return ""
        
        try:
            h_sutun_idx = VERI_SUTUN_INDEKSLERI['ILGILI_KESINTILER']
            
            for idx, row in self.df_analiz.iterrows():
                analiz_grup = str(row.iloc[h_sutun_idx]).strip()
                analiz_grup_normalized = ExcelYardimci.normalize_grup_string(analiz_grup)
                
                if analiz_grup_normalized == grup:
                    # "Ortak W Değerleri" sütununu bul
                    if 'Ortak W Değerleri' in self.df_analiz.columns:
                        return str(row['Ortak W Değerleri'])
                    # Alternatif: Son sütun
                    return str(row.iloc[-1]) if len(row) > 0 else ""
            
            return ""
        except Exception as e:
            print(f"W değeri alma hatası: {e}")
            return ""
    
    def _filter_table_columns(self, df):
        """table.xlsx için sütun filtresi"""
        try:
            selected_indices = TABLE_SUTUN_INDEKSLERI['SECILEN_SUTUNLAR']
            new_df = df.iloc[:, selected_indices].copy()
            
            # Toplam etkilenen kullanıcı hesapla
            try:
                t_col = pd.to_numeric(df.iloc[:, TABLE_SUTUN_INDEKSLERI['ETKILENEN_KULLANICI_T']], errors='coerce').fillna(0)
                u_col = pd.to_numeric(df.iloc[:, TABLE_SUTUN_INDEKSLERI['ETKILENEN_KULLANICI_U']], errors='coerce').fillna(0)
                v_col = pd.to_numeric(df.iloc[:, TABLE_SUTUN_INDEKSLERI['ETKILENEN_KULLANICI_V']], errors='coerce').fillna(0)
                w_col = pd.to_numeric(df.iloc[:, TABLE_SUTUN_INDEKSLERI['ETKILENEN_KULLANICI_W']], errors='coerce').fillna(0)
                new_df['Toplam Etkilenen Kullanıcı'] = (t_col + u_col + v_col + w_col).astype(int)
            except:
                new_df['Toplam Etkilenen Kullanıcı'] = 0
            
            # Ek sütunlar
            for idx in TABLE_SUTUN_INDEKSLERI['EK_SUTUNLAR']:
                if idx < len(df.columns):
                    new_df[df.columns[idx]] = df.iloc[:, idx]
            
            # OMS yorum
            oms_idx = TABLE_SUTUN_INDEKSLERI['OMS_YORUM']
            if oms_idx < len(df.columns):
                new_df[df.columns[oms_idx]] = df.iloc[:, oms_idx]
            
            # İlk çağrı zamanı
            try:
                at_idx = TABLE_SUTUN_INDEKSLERI['ILK_CAGRI_AT']
                au_idx = TABLE_SUTUN_INDEKSLERI['ILK_CAGRI_AU']
                at_col = pd.to_datetime(df.iloc[:, at_idx], errors='coerce')
                au_col = pd.to_datetime(df.iloc[:, au_idx], errors='coerce')
                
                min_times = []
                for at_val, au_val in zip(at_col, au_col):
                    at_valid = at_val is not pd.NaT
                    au_valid = au_val is not pd.NaT
                    
                    if at_valid and au_valid:
                        if at_val <= au_val:
                            min_times.append(f"{at_val.strftime('%d.%m.%Y %H:%M:%S')} (Müşteri Dışı)")
                        else:
                            min_times.append(f"{au_val.strftime('%d.%m.%Y %H:%M:%S')} (Müşteri)")
                    elif at_valid:
                        min_times.append(f"{at_val.strftime('%d.%m.%Y %H:%M:%S')} (Müşteri Dışı)")
                    elif au_valid:
                        min_times.append(f"{au_val.strftime('%d.%m.%Y %H:%M:%S')} (Müşteri)")
                    else:
                        min_times.append("")
                new_df['İlk Çağrı Zamanı'] = min_times
            except:
                new_df['İlk Çağrı Zamanı'] = ""
            
            # Son sütun
            son_idx = TABLE_SUTUN_INDEKSLERI['SON_SUTUN']
            if son_idx < len(df.columns):
                new_df[df.columns[son_idx]] = df.iloc[:, son_idx]
            
            return new_df
        except Exception as e:
            print(f"Filtreleme hatası: {e}")
            return df
    
    def png_olustur(self, id_listesi, grup_adi, grup_folder, kaynak_adi):
        """
        PNG raporu oluştur (optimize edilmiş versiyon).
        
        Args:
            id_listesi: Kesinti ID listesi
            grup_adi: Grup adı
            grup_folder: Hedef klasör
            kaynak_adi: 'OTG' veya 'JTK'
        """
        if kaynak_adi == 'OTG':
            df = self.df_table
            col_index = TABLE_SUTUN_INDEKSLERI['KESINTI_ID']
            is_table = True
        else:
            df = self.df_jtk
            col_index = JTK_SUTUN_INDEKSLERI['KESINTI_ID']
            is_table = False
        
        if df is None:
            return
        
        all_data = []
        for aranan_id in id_listesi:
            data = ExcelYardimci.id_ara(df, aranan_id, col_index)
            
            if is_table and len(data) > 0:
                data = self._filter_table_columns(data)
            
            all_data.append({'id': aranan_id, 'data': data})
        
        total_rows = sum(1 + len(item['data']) for item in all_data if len(item['data']) > 0)
        if total_rows == 0:
            return
        
        # Önce veriyi hazırla (fig boyutunu hesaplamak için)
        combined_data = []
        cell_line_counts = []  # Her satırdaki maksimum satır sayısı
        
        # OMS sütun indeksini önceden bul
        oms_col_idx = -1
        if kaynak_adi == 'OTG' and len(all_data) > 0:
            first_item = next((item for item in all_data if len(item['data']) > 0), None)
            if first_item:
                for idx, col_name in enumerate(first_item['data'].columns):
                    if 'OMS Kesinti Yorumu' in str(col_name):
                        oms_col_idx = idx
                        break
        
        # Wrap uzunlukları - OTG ve JTK için farklı
        if kaynak_adi == 'OTG':
            header_wrap_len = PNG_AYARLARI.get('HEADER_WRAP_LENGTH', 15)
            cell_wrap_len = PNG_AYARLARI.get('OTG_CELL_WRAP', 18)
            oms_wrap_len = 40
        else:
            header_wrap_len = PNG_AYARLARI.get('HEADER_WRAP_LENGTH', 20)
            cell_wrap_len = PNG_AYARLARI.get('CELL_WRAP_LENGTH', 25)
            oms_wrap_len = 50
        
        for item in all_data:
            if len(item['data']) > 0:
                # Başlıklar - HER ZAMAN wrap
                headers = []
                header_max_lines = 1
                for col in item['data'].columns:
                    col_str = str(col)
                    wrapped = self._wrap_text(col_str, header_wrap_len)
                    headers.append(wrapped)
                    header_max_lines = max(header_max_lines, wrapped.count('\n') + 1)
                combined_data.append(headers)
                cell_line_counts.append(header_max_lines)
                
                # Veri satırları
                for _, row in item['data'].iterrows():
                    row_data = []
                    max_lines_in_row = 1
                    
                    for col_idx, val in enumerate(row.values):
                        # Değeri temizle ve formatla
                        formatted_val = ExcelYardimci.temizle_ve_formatla(val, max_karakter=None, wrap_satir=False)
                        
                        # Manuel wrap uygula
                        if col_idx == oms_col_idx:
                            wrapped = self._wrap_text(formatted_val, oms_wrap_len)
                        else:
                            wrapped = self._wrap_text(formatted_val, cell_wrap_len)
                        
                        row_data.append(wrapped)
                        max_lines_in_row = max(max_lines_in_row, wrapped.count('\n') + 1)
                    
                    combined_data.append(row_data)
                    cell_line_counts.append(max_lines_in_row)
        
        if len(combined_data) == 0:
            return
        
        # Figure boyutunu içeriğe göre hesapla
        total_lines = sum(cell_line_counts)
        row_height_factor = PNG_AYARLARI.get('ROW_HEIGHT_FACTOR', 0.4)
        fig_height = max(total_lines * row_height_factor, 4.0)
        fig_width = PNG_AYARLARI['OTG_FIG_WIDTH'] if kaynak_adi == 'OTG' else PNG_AYARLARI['JTK_FIG_WIDTH']
        
        fig, ax = plt.subplots(1, 1, figsize=(fig_width, fig_height))
        ax.axis('off')
        
        # Başlık yok - sadece tablo
        
        # Tabloyu oluştur
        table = ax.table(cellText=combined_data, cellLoc='center', loc='center')
        
        font_size = PNG_AYARLARI['OTG_FONT_SIZE'] if kaynak_adi == 'OTG' else PNG_AYARLARI['JTK_FONT_SIZE']
        header_font_size = PNG_AYARLARI.get('HEADER_FONT_SIZE', font_size + 1)
        
        table.auto_set_font_size(False)
        table.set_fontsize(font_size)
        
        # Tabloyu ölçeklendir (satır sayısına ve kaynak tipine göre)
        # Büyük yazı için daha fazla alan
        if kaynak_adi == 'OTG':
            table.scale(1.3, 2.2 + (total_lines * 0.12))
        else:
            table.scale(1.1, 2.0 + (total_lines * 0.1))
        
        # Sütun genişliklerini hesapla
        num_cols = len(combined_data[0])
        col_widths = []
        
        for col_idx in range(num_cols):
            max_line_len = 0
            for row in combined_data:
                if col_idx < len(row):
                    cell_text = str(row[col_idx])
                    if cell_text:
                        for line in cell_text.split('\n'):
                            max_line_len = max(max_line_len, len(line.strip()))
            col_widths.append(max(5, max_line_len + 2))
        
        total_width = sum(col_widths)
        
        # Her hücreyi ayarla
        for row_idx in range(len(combined_data)):
            for col_idx in range(num_cols):
                if col_idx < len(combined_data[row_idx]):
                    cell = table[(row_idx, col_idx)]
                    # Genişlik
                    cell.set_width(col_widths[col_idx] / total_width)
                    # Yükseklik (satır sayısına göre) - BÜYÜK YAZI için artırıldı
                    line_count = cell_line_counts[row_idx] if row_idx < len(cell_line_counts) else 1
                    cell.set_height(0.07 + (line_count - 1) * 0.04)
        
        # Stil uygula
        row_idx = 0
        header_color = PNG_AYARLARI['HEADER_COLOR']
        alt_color = PNG_AYARLARI['ALTERNATE_ROW_COLOR']
        
        for item in all_data:
            if len(item['data']) > 0:
                # Başlık satırı
                for col_idx in range(num_cols):
                    cell = table[(row_idx, col_idx)]
                    cell.set_facecolor(header_color)
                    cell.set_text_props(
                        weight='bold', 
                        color='white',
                        fontsize=header_font_size
                    )
                
                row_idx += 1
                
                # Veri satırları
                for data_row_idx in range(len(item['data'])):
                    for col_idx in range(num_cols):
                        cell = table[(row_idx, col_idx)]
                        if data_row_idx % 2 == 0:
                            cell.set_facecolor(alt_color)
                    row_idx += 1
        
        plt.tight_layout(pad=0.2)
        
        png_path = os.path.join(grup_folder, f"{grup_adi}-{kaynak_adi}.png")
        plt.savefig(png_path, dpi=PNG_AYARLARI['DPI'], bbox_inches='tight',
                    facecolor='none', edgecolor='none', format='png', 
                    pad_inches=0.05, transparent=True)
        plt.close(fig)
        print(f"  ✓ {kaynak_adi}.png oluşturuldu")
    
    def _wrap_text(self, text, max_len):
        """Metni belirli uzunlukta satırlara böl"""
        if not text or len(str(text)) <= max_len:
            return str(text) if text else ''
        
        text = str(text)
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            # Kelime çok uzunsa böl
            while len(word) > max_len:
                if current_line:
                    lines.append(current_line)
                    current_line = ""
                lines.append(word[:max_len])
                word = word[max_len:]
            
            if not word:
                continue
                
            if not current_line:
                current_line = word
            elif len(current_line + " " + word) <= max_len:
                current_line += " " + word
            else:
                lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        return "\n".join(lines) if lines else str(text)
    
    def cm_excel_olustur(self, id_listesi, grup_adi, grup_folder):
        """
        CM Excel raporu oluştur.
        
        Args:
            id_listesi: Kesinti ID listesi
            grup_adi: Grup adı
            grup_folder: Hedef klasör
        """
        if self.df_cm is None:
            return
        
        col_index = CM_SUTUN_INDEKSLERI['KESINTI_ID']
        
        all_data = []
        for aranan_id in id_listesi:
            data = ExcelYardimci.id_ara(self.df_cm, aranan_id, col_index)
            if len(data) > 0:
                all_data.append({'id': aranan_id, 'data': data})
        
        if len(all_data) == 0:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"CM_{grup_adi[:20]}"
        
        header_fill = PatternFill(start_color=EXCEL_STIL['HEADER_BG_COLOR'], 
                                   end_color=EXCEL_STIL['HEADER_BG_COLOR'], 
                                   fill_type="solid")
        header_font = Font(bold=True, color=EXCEL_STIL['HEADER_FONT_COLOR'])
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_fill_even = PatternFill(start_color=EXCEL_STIL['ALTERNATE_ROW_COLOR'], 
                                      end_color=EXCEL_STIL['ALTERNATE_ROW_COLOR'], 
                                      fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        row_idx = 1
        for item in all_data:
            # Başlık
            for col_idx, col_name in enumerate(item['data'].columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            row_idx += 1
            
            # Veriler
            for _, data_row in item['data'].iterrows():
                for col_idx, value in enumerate(data_row.values, 1):
                    cell_value = ExcelYardimci.format_tarih(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.alignment = data_alignment
                    if (row_idx - 1) % 2 == 0:
                        cell.fill = data_fill_even
                row_idx += 1
        
        # Sütun genişlikleri
        max_width = EXCEL_STIL.get('MAX_COLUMN_WIDTH', 50)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column].width = adjusted_width
        
        excel_path = os.path.join(grup_folder, f"{grup_adi}-CM.xlsx")
        wb.save(excel_path)
        print(f"  ✓ CM.xlsx oluşturuldu")
    
    def _grup_kaynaga_gore_al(self, grup):
        """
        Analiz sonucundan grubun "Kaynağa Göre" bilgisini al.
        
        Args:
            grup: Grup string'i (;'li kesinti ID'leri)
            
        Returns:
            str: Kaynağa Göre değeri veya boş string
        """
        if self.df_analiz is None:
            return ""
        
        try:
            h_sutun_idx = VERI_SUTUN_INDEKSLERI['ILGILI_KESINTILER']
            
            for idx, row in self.df_analiz.iterrows():
                analiz_grup = str(row.iloc[h_sutun_idx]).strip()
                analiz_grup_normalized = ExcelYardimci.normalize_grup_string(analiz_grup)
                
                if analiz_grup_normalized == grup:
                    # "Kaynağa Göre" sütununu bul
                    if 'Kaynağa Göre' in self.df_analiz.columns:
                        return str(row['Kaynağa Göre']).strip()
            
            return ""
        except Exception as e:
            print(f"Kaynağa Göre alma hatası: {e}")
            return ""
    
    def tum_gruplari_isle(self, progress_callback=None):
        """
        Tüm grupları işle.
        
        Args:
            progress_callback: İlerleme callback fonksiyonu (idx, total, grup)
            
        Returns:
            int: İşlenen grup sayısı
        """
        output_base = os.path.join(self.klasor_yolu, VARSAYILAN['OUTPUT_FOLDER'])
        os.makedirs(output_base, exist_ok=True)
        
        # Dağıtım-AG ayarları
        dagitim_ag_deger = DAGITIM_AG_AYARLARI.get('KAYNAGA_GORE_DEGER', 'Dağıtım-AG')
        jtk_olustur = DAGITIM_AG_AYARLARI.get('JTK_OLUSTUR', False)
        
        for idx, grup in enumerate(self.grup_list, 1):
            print(f"\nGRUP {idx}/{len(self.grup_list)}: {grup}")
            
            if progress_callback:
                progress_callback(idx, len(self.grup_list), grup)
            
            grup_folder = os.path.join(output_base, grup)
            os.makedirs(grup_folder, exist_ok=True)
            
            id_listesi = [id.strip() for id in grup.split(';')]
            
            # Kaynağa Göre bilgisini al
            kaynaga_gore = self._grup_kaynaga_gore_al(grup)
            is_dagitim_ag = (kaynaga_gore == dagitim_ag_deger)
            
            # OTG PNG'sini oluştur (her zaman)
            self.png_olustur(id_listesi, grup, grup_folder, 'OTG')
            
            # JTK PNG'sini oluştur (Dağıtım-AG değilse veya JTK_OLUSTUR=True ise)
            if not is_dagitim_ag or jtk_olustur:
                self.png_olustur(id_listesi, grup, grup_folder, 'JTK')
            else:
                print(f"  ⏭️ JTK atlandı (Dağıtım-AG)")
            
            # CM Excel oluştur
            self.cm_excel_olustur(id_listesi, grup, grup_folder)
        
        return len(self.grup_list)

